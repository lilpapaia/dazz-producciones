from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from typing import List
from io import BytesIO

from sqlalchemy.orm import Session


def _resolve_billing_company_name(project, db: Session) -> str:
    """Resolve billing company name from oc_prefixes via project creative_code."""
    if not db or not project.creative_code:
        return getattr(project.owner_company, 'name', '') or project.company or ''

    try:
        from app.models.database import OCPrefix
        from sqlalchemy import func

        oc_upper = project.creative_code.strip().upper()
        prefixes = db.query(OCPrefix).filter(
            OCPrefix.active == True
        ).order_by(func.length(OCPrefix.prefix).desc()).all()

        for p in prefixes:
            if oc_upper.startswith(p.prefix.upper()):
                if p.billing_company:
                    return p.billing_company.name
                break
    except Exception:
        pass

    return getattr(project.owner_company, 'name', '') or project.company or ''


def create_project_excel_bytes(project, tickets, db: Session = None) -> bytes:
    """
    Create Excel in memory (BytesIO) and return bytes.
    Matches the DESGLOSE_CON_MARGEN format used by accounting.

    Args:
        project: SQLAlchemy Project object
        tickets: List of SQLAlchemy Ticket objects
        db: SQLAlchemy Session (for resolving billing company)

    Returns:
        bytes: Excel file as bytes
    """

    wb = Workbook()
    sheet = wb.active
    sheet.title = "Hoja 1"

    # ── Styles ──────────────────────────────────────────────
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    style_header_meta = {
        'font': Font(name='Arial', size=10, color='000000'),
        'fill': PatternFill(start_color='ED95DF', end_color='ED95DF', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'border': thin_border,
    }

    style_data_row = {
        'font': Font(name='Calibri', size=11),
        'fill': PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'border': thin_border,
    }

    style_header_tickets = {
        'font': Font(name='Calibri', size=11, bold=True, color='FFFFFF'),
        'fill': PatternFill(start_color='000000', end_color='000000', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'border': thin_border,
    }

    style_ticket = {
        'font': Font(name='Calibri', size=11),
        'alignment': Alignment(horizontal='center', vertical='center'),
        'border': thin_border,
    }

    style_footer = {
        'font': Font(name='Calibri', size=11, bold=True),
        'fill': PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center'),
        'border': thin_border,
    }

    style_footer_diff = {
        'font': Font(name='Calibri', size=11, bold=True),
        'fill': PatternFill(start_color='AEAAAA', end_color='AEAAAA', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center'),
        'border': thin_border,
    }

    def _apply(cell, style: dict):
        for attr, val in style.items():
            setattr(cell, attr, val)

    # ── Resolve billing company name ────────────────────────
    company_name = getattr(project.owner_company, 'name', '') or project.company or ''
    billing_company_name = _resolve_billing_company_name(project, db)

    # ── FILA 1: Cabecera metadatos (19 cols, rosa) ──────────
    headers_meta = [
        "TIPO", "AÑO", "FECHA ENVIO FACTURAR", "OC PROYECTO",
        "EMPRESA FACTURACION", "QUIEN PASA FACTURAR (TAG)", "TIPO FACTURA",
        "QUIEN INTERVIENE", "DESCRIPCIÓN/CAMPAÑA/ACCIÓN", "IMPORTE BRUTO TOTAL",
        "MONEDA", "IRPF", "IMPORTE IRPF", "OC DE CLIENTE",
        "OTROS DATOS FACTURA", "DATOS CLIENTE", "EMAIL CLIENTE",
        "ESTATUS", "LINEA FISPER",
    ]
    for col, header in enumerate(headers_meta, start=1):
        cell = sheet.cell(row=1, column=col, value=header)
        _apply(cell, style_header_meta)

    # ── FILA 2: Datos del proyecto ──────────────────────────
    # Pre-calculate ticket data row range for SUM formula
    ticket_start_row = 4
    ticket_end_row = ticket_start_row + len(tickets) - 1 if tickets else ticket_start_row

    project_data = [
        company_name,                                                    # A: TIPO
        int(project.year) if project.year else '',                       # B: AÑO
        project.send_date or '',                                         # C: FECHA ENVIO FACTURAR
        project.creative_code or '',                                     # D: OC PROYECTO
        billing_company_name,                                            # E: EMPRESA FACTURACION
        project.responsible or '',                                       # F: QUIEN PASA FACTURAR
        project.invoice_type or '',                                      # G: TIPO FACTURA
        '',                                                              # H: QUIEN INTERVIENE (vacío)
        project.description or '',                                       # I: DESCRIPCIÓN/CAMPAÑA
        f'=SUM(E{ticket_start_row}:E{ticket_end_row})' if tickets else 0,  # J: IMPORTE BRUTO TOTAL
        '',                                                              # K: MONEDA (vacío)
        '',                                                              # L: IRPF (vacío)
        '',                                                              # M: IMPORTE IRPF (vacío)
        project.client_oc or '',                                         # N: OC DE CLIENTE
        project.other_invoice_data or '',                                # O: OTROS DATOS FACTURA
        project.client_data or '',                                       # P: DATOS CLIENTE
        project.client_email or '',                                      # Q: EMAIL CLIENTE
        project.status.value if hasattr(project.status, 'value') else str(project.status or ''),  # R: ESTATUS
        '',                                                              # S: LINEA FISPER (vacío)
    ]
    for col, value in enumerate(project_data, start=1):
        cell = sheet.cell(row=2, column=col, value=value)
        _apply(cell, style_data_row)
    # Apply currency format to IMPORTE BRUTO TOTAL
    sheet.cell(row=2, column=10).number_format = '#,##0.00 €'

    # ── FILA 3: Cabecera gastos (22 cols, negro) ────────────
    headers_tickets = [
        "FECHA FACTURA", "Proveedor - Nombre", "Nº FACTURA PROVEEDOR",
        "DESCRIPCIÓN", "IMPORTE", "TIPO", "TIPO IVA", "TOTAL",
        "TIPO IRPF", "RETENCION", "TOTAL",
        "ESTATUS FACTURA", "ESTATUS PAGO",
        "TELEFONO", "EMAIL", "NOMBRE",
        "ESTATUS EN CONTABILIDAD", "nº de GASTO", "ESTATUS PAGO",
        "LINK", "COMO SE PAGÓ", "FECHA PAGO",
    ]
    for col, header in enumerate(headers_tickets, start=1):
        cell = sheet.cell(row=3, column=col, value=header)
        _apply(cell, style_header_tickets)

    # ── FILAS 4+: Tickets ───────────────────────────────────
    for idx, ticket in enumerate(tickets, start=ticket_start_row):
        row_data = [
            ticket.date or '',                                          # A: FECHA FACTURA
            ticket.provider or '',                                      # B: Proveedor
            ticket.invoice_number or '',                                # C: Nº FACTURA
            getattr(ticket, 'notes', '') or '',                         # D: DESCRIPCIÓN
            ticket.base_amount or 0,                                    # E: IMPORTE
            ticket.type.value if hasattr(ticket.type, 'value') else str(ticket.type or ''),  # F: TIPO
            (ticket.iva_percentage or 0) / 100 if ticket.iva_percentage else 0,  # G: TIPO IVA (%)
            (ticket.base_amount or 0) + (ticket.iva_amount or 0),       # H: TOTAL (base+iva)
            (getattr(ticket, 'irpf_percentage', 0) or 0) / 100 if getattr(ticket, 'irpf_percentage', 0) else 0,  # I: TIPO IRPF (%)
            getattr(ticket, 'irpf_amount', 0) or 0,                    # J: RETENCION
            ticket.final_total or 0,                                    # K: TOTAL
            ticket.invoice_status or '',                                # L: ESTATUS FACTURA
            ticket.payment_status or '',                                # M: ESTATUS PAGO
        ]

        for col, value in enumerate(row_data, start=1):
            cell = sheet.cell(row=idx, column=col, value=value)
            _apply(cell, style_ticket)

        # Contact info (cols N-P, only for facturas)
        if ticket.type == 'factura' or (hasattr(ticket.type, 'value') and ticket.type.value == 'factura'):
            sheet.cell(row=idx, column=14, value=ticket.phone or '')
            sheet.cell(row=idx, column=15, value=ticket.email or '')
            sheet.cell(row=idx, column=16, value=getattr(ticket, 'contact_name', '') or '')

        # Apply border to remaining empty cells (cols N-V)
        for col in range(14, 23):
            cell = sheet.cell(row=idx, column=col)
            cell.border = thin_border
            cell.alignment = style_ticket['alignment']
            cell.font = style_ticket['font']

        # Number formats
        sheet.cell(row=idx, column=5).number_format = '#,##0.00 €'    # IMPORTE
        sheet.cell(row=idx, column=7).number_format = '0%'            # TIPO IVA
        sheet.cell(row=idx, column=8).number_format = '#,##0.00 €'    # TOTAL base+iva
        sheet.cell(row=idx, column=9).number_format = '0%'            # TIPO IRPF
        sheet.cell(row=idx, column=10).number_format = '#,##0.00 €'   # RETENCION
        sheet.cell(row=idx, column=11).number_format = '#,##0.00 €'   # TOTAL final

    # ── FOOTER ──────────────────────────────────────────────
    footer_row = ticket_start_row + len(tickets)
    num_ticket_cols = len(headers_tickets)

    # Row: TOTAL
    for col in range(1, num_ticket_cols + 1):
        cell = sheet.cell(row=footer_row, column=col)
        _apply(cell, style_footer)
    sheet.cell(row=footer_row, column=5, value="TOTAL")
    sheet.cell(row=footer_row, column=5).font = Font(name='Calibri', size=11, bold=True)
    if tickets:
        sheet.cell(row=footer_row, column=6, value=f'=SUM(E{ticket_start_row}:E{ticket_end_row})')
    else:
        sheet.cell(row=footer_row, column=6, value=0)
    sheet.cell(row=footer_row, column=6).number_format = '#,##0.00 €'

    # Row: PRESUPUESTO
    budget_row = footer_row + 1
    for col in range(1, num_ticket_cols + 1):
        cell = sheet.cell(row=budget_row, column=col)
        _apply(cell, style_footer)
    sheet.cell(row=budget_row, column=5, value="PRESUPUESTO")
    sheet.cell(row=budget_row, column=5).font = Font(name='Calibri', size=11, bold=True)
    presupuesto = getattr(project, 'presupuesto', None)
    sheet.cell(row=budget_row, column=6, value=presupuesto if presupuesto is not None else '')
    sheet.cell(row=budget_row, column=6).number_format = '#,##0.00 €'

    # Row: DIFERENCIA
    diff_row = footer_row + 2
    for col in range(1, num_ticket_cols + 1):
        cell = sheet.cell(row=diff_row, column=col)
        _apply(cell, style_footer_diff)
    sheet.cell(row=diff_row, column=5, value="DIFERENCIA")
    sheet.cell(row=diff_row, column=5).font = Font(name='Calibri', size=11, bold=True)
    if presupuesto is not None:
        sheet.cell(row=diff_row, column=6, value=f'=F{budget_row}-F{footer_row}')
    else:
        sheet.cell(row=diff_row, column=6, value='')
    sheet.cell(row=diff_row, column=6).number_format = '#,##0.00 €'

    # Row: MARGEN %
    margin_row = footer_row + 3
    for col in range(1, num_ticket_cols + 1):
        cell = sheet.cell(row=margin_row, column=col)
        _apply(cell, style_footer)
    sheet.cell(row=margin_row, column=5, value="PORCENTAJE MARGEN")
    sheet.cell(row=margin_row, column=5).font = Font(name='Calibri', size=11, bold=True)
    if presupuesto is not None:
        sheet.cell(row=margin_row, column=6, value=f'=IF(F{budget_row}=0,0,F{diff_row}/F{budget_row})')
    else:
        sheet.cell(row=margin_row, column=6, value='')
    sheet.cell(row=margin_row, column=6).number_format = '0%'

    # ── LINK A PROYECTO (fila extra) ────────────────────────
    if project.project_link:
        link_row = margin_row + 2
        cell_label = sheet.cell(row=link_row, column=1, value="LINK A PROYECTO:")
        cell_label.font = Font(name='Calibri', size=11, bold=True)
        cell_link = sheet.cell(row=link_row, column=2, value=project.project_link)
        cell_link.font = Font(name='Calibri', size=11, color='0563C1', underline='single')

    # ── Column widths: all 13 ───────────────────────────────
    max_cols = max(len(headers_meta), len(headers_tickets))
    for i in range(1, max_cols + 1):
        col_letter = sheet.cell(row=1, column=i).column_letter
        sheet.column_dimensions[col_letter].width = 13

    # ── Save to memory ──────────────────────────────────────
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)

    return excel_buffer.getvalue()
