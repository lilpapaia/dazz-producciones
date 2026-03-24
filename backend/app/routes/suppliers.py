"""
Endpoints de gestión de proveedores — Solo ADMIN.
Prefijo: /suppliers
"""

import logging
from fastapi import APIRouter, Body, Depends, HTTPException, status, Query, Request

logger = logging.getLogger(__name__)
from sqlalchemy.orm import Session, joinedload
from sqlalchemy import func, desc, case
from datetime import datetime, timedelta, timezone
from typing import Optional, List
import secrets
import string

from config.database import get_db
from app.models.database import User, Company, Project, ProjectStatus, Ticket, TicketType
from app.models.suppliers import (
    Supplier, SupplierInvoice, SupplierOC, SupplierNotification,
    SupplierInvitation, InvoiceStatus, SupplierStatus,
    NotificationRecipientType, NotificationEventType,
)
from app.services.auth import get_current_admin_user
from app.models.supplier_schemas import (
    InviteRequest, InviteResponse, SupplierResponse, SupplierUpdate,
    AssignOCRequest, AssignInvoiceOCRequest, NoteRequest, InvoiceStatusUpdate,
    InvoiceResponse, InvoiceDetailResponse,
    NotificationResponse, DashboardResponse, CreateOCRequest, CreateOCResponse,
)
from app.services.supplier_auth import invalidate_all_supplier_tokens
from app.services.supplier_storage import get_invoice_pdf_url, get_bank_cert_url, delete_invoice_pdf
from app.services.cloudinary_service import extract_public_id_from_url
from app.services.encryption import decrypt_iban, encrypt_iban, is_encryption_available
from app.services.supplier_ai import format_date_for_response, parse_invoice_date
from app.services.supplier_email import (
    send_supplier_invitation,
    send_supplier_invoice_paid,
)

router = APIRouter(prefix="/suppliers", tags=["Suppliers (Admin)"])


def _decode_iban(supplier) -> str | None:
    """Decrypt IBAN for admin view (full, unmasked)."""
    if not supplier.iban_encrypted:
        return None
    return decrypt_iban(supplier.iban_encrypted)


def _generate_token(length: int = 64) -> str:
    chars = string.ascii_letters + string.digits
    return "".join(secrets.choice(chars) for _ in range(length))


def _notify(db: Session, recipient_type, recipient_id: int, event_type,
            title: str, message: str, invoice_id=None, supplier_id=None):
    """Helper to create a notification."""
    notif = SupplierNotification(
        recipient_type=recipient_type,
        recipient_id=recipient_id,
        event_type=event_type,
        title=title,
        message=message,
        related_invoice_id=invoice_id,
        related_supplier_id=supplier_id,
    )
    db.add(notif)


# ============================================
# OC MANAGEMENT
# ============================================

@router.post("/ocs", response_model=CreateOCResponse, status_code=201)
async def create_oc(
    body: CreateOCRequest,
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin_user),
):
    """Create a new OC for a talent/influencer."""
    existing = db.query(SupplierOC).filter(SupplierOC.oc_number == body.oc_number.strip()).first()
    if existing:
        raise HTTPException(400, f"OC {body.oc_number} already exists")

    if body.company_id:
        company = db.query(Company).filter(Company.id == body.company_id).first()
        if not company:
            raise HTTPException(404, "Company not found")

    oc = SupplierOC(
        oc_number=body.oc_number.strip(),
        talent_name=body.talent_name.strip(),
        nif_cif=body.nif_cif.strip() if body.nif_cif else None,
        company_id=body.company_id,
    )
    db.add(oc)
    db.commit()
    db.refresh(oc)

    return CreateOCResponse(
        id=oc.id,
        oc_number=oc.oc_number,
        talent_name=oc.talent_name,
        nif_cif=oc.nif_cif,
        company_id=oc.company_id,
    )


@router.get("/oc-suggestions")
async def oc_suggestions(
    q: str = Query("", max_length=50),
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin_user),
):
    """Autocomplete: search permanent OCs + open projects by code/name."""
    results = []
    term = q.strip()
    if len(term) < 2:
        return results

    # 1. OCs permanentes (supplier_ocs)
    ocs = db.query(SupplierOC).filter(
        SupplierOC.oc_number.ilike(f"%{term}%")
    ).limit(5).all()
    for oc in ocs:
        company_name = None
        if oc.company_id:
            co = db.query(Company).filter(Company.id == oc.company_id).first()
            company_name = co.name if co else None
        results.append({
            "type": "permanent_oc",
            "oc_number": oc.oc_number,
            "label": oc.talent_name,
            "company_name": company_name,
        })

    # 2. Proyectos abiertos (creative_code o description)
    projects = db.query(Project).filter(
        Project.status == ProjectStatus.EN_CURSO,
        (Project.creative_code.ilike(f"%{term}%") | Project.description.ilike(f"%{term}%")),
    ).limit(5).all()
    for p in projects:
        company_name = None
        if p.owner_company_id:
            co = db.query(Company).filter(Company.id == p.owner_company_id).first()
            company_name = co.name if co else None
        results.append({
            "type": "project",
            "oc_number": p.creative_code,
            "label": p.description or p.creative_code,
            "company_name": company_name,
            "project_id": p.id,
            "company_id": p.owner_company_id,
        })

    return results


# ============================================
# INVITATIONS
# ============================================

@router.post("/invite", response_model=InviteResponse, status_code=201)
async def invite_supplier(
    body: InviteRequest,
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin_user),
):
    """Send invitation email to a new supplier."""
    # Check if supplier already exists
    existing = db.query(Supplier).filter(Supplier.email == body.email).first()
    if existing:
        raise HTTPException(400, "A supplier with this email already exists")

    # Check for pending invitation
    pending = db.query(SupplierInvitation).filter(
        SupplierInvitation.email == body.email,
        SupplierInvitation.used_at == None,
        SupplierInvitation.expires_at > datetime.now(timezone.utc),
    ).first()
    if pending:
        raise HTTPException(400, "An active invitation already exists for this email")

    token = _generate_token()
    expires_at = datetime.now(timezone.utc) + timedelta(hours=72)

    invitation = SupplierInvitation(
        email=body.email,
        name=body.name,
        token=token,
        expires_at=expires_at,
        invited_by=admin.id,
    )
    db.add(invitation)
    db.commit()
    db.refresh(invitation)

    # Send email (non-blocking — don't fail if email fails)
    try:
        send_supplier_invitation(body.name, body.email, token, body.message)
    except Exception as e:
        logger.warning(f"Invitation email failed: {e}")

    return InviteResponse(
        id=invitation.id,
        name=body.name,
        email=body.email,
        expires_at=expires_at,
        message="Invitation sent successfully",
    )


# ============================================
# SUPPLIER CRUD
# ============================================

@router.get("", response_model=List[SupplierResponse])
async def list_suppliers(
    status_filter: Optional[str] = Query(None, alias="status"),
    company_id: Optional[int] = None,
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin_user),
):
    """List all suppliers with optional filters."""
    query = db.query(Supplier).options(
        joinedload(Supplier.oc).joinedload(SupplierOC.company)
    )

    if status_filter:
        query = query.filter(Supplier.status == status_filter)

    if company_id:
        oc_ids = [oc.id for oc in db.query(SupplierOC).filter(SupplierOC.company_id == company_id).all()]
        query = query.filter(Supplier.oc_id.in_(oc_ids)) if oc_ids else query.filter(False)

    suppliers = query.order_by(desc(Supplier.created_at)).all()

    if not suppliers:
        return []

    # Batch pre-fetch: invoice counts (total + pending) per supplier
    supplier_ids = [s.id for s in suppliers]

    counts = db.query(
        SupplierInvoice.supplier_id,
        func.count(SupplierInvoice.id).label("total"),
        func.count(case((SupplierInvoice.status == InvoiceStatus.PENDING, 1))).label("pending"),
    ).filter(
        SupplierInvoice.supplier_id.in_(supplier_ids)
    ).group_by(SupplierInvoice.supplier_id).all()
    counts_map = {c.supplier_id: (c.total, c.pending) for c in counts}

    # Batch pre-fetch: last activity timestamps
    last_inv_map = dict(db.query(
        SupplierInvoice.supplier_id,
        func.max(SupplierInvoice.created_at),
    ).filter(
        SupplierInvoice.supplier_id.in_(supplier_ids)
    ).group_by(SupplierInvoice.supplier_id).all())

    last_notif_map = dict(db.query(
        SupplierNotification.related_supplier_id,
        func.max(SupplierNotification.created_at),
    ).filter(
        SupplierNotification.related_supplier_id.in_(supplier_ids)
    ).group_by(SupplierNotification.related_supplier_id).all())

    # Build responses using pre-loaded data (no per-supplier queries)
    result = []
    for s in suppliers:
        oc_number = s.oc.oc_number if s.oc else None
        company_name = s.oc.company.name if s.oc and s.oc.company else None
        inv_total, inv_pending = counts_map.get(s.id, (0, 0))
        last_activity = max(filter(None, [
            last_notif_map.get(s.id), last_inv_map.get(s.id), s.created_at
        ]))

        result.append(SupplierResponse(
            id=s.id, name=s.name, email=s.email, nif_cif=s.nif_cif,
            phone=s.phone, address=s.address, iban=_decode_iban(s),
            bank_cert_url=s.bank_cert_url,
            status=s.status.value if s.status else "NEW",
            has_permanent_oc=s.oc_id is not None,
            oc_id=s.oc_id, oc_number=oc_number, company_name=company_name,
            is_active=s.is_active, notes_internal=s.notes_internal,
            gdpr_consent=s.gdpr_consent,
            created_at=s.created_at, updated_at=s.updated_at,
            last_activity=last_activity,
            invoices_count=inv_total, pending_invoices=inv_pending,
        ))
    return result


def _build_supplier_response(s: Supplier, db: Session) -> SupplierResponse:
    """Build a full SupplierResponse with computed fields."""
    inv_count = db.query(func.count(SupplierInvoice.id)).filter(
        SupplierInvoice.supplier_id == s.id).scalar()
    pending = db.query(func.count(SupplierInvoice.id)).filter(
        SupplierInvoice.supplier_id == s.id,
        SupplierInvoice.status == InvoiceStatus.PENDING,
    ).scalar()

    oc_number = None
    company_name = None
    if s.oc_id:
        oc = db.query(SupplierOC).filter(SupplierOC.id == s.oc_id).first()
        if oc:
            oc_number = oc.oc_number
            if oc.company_id:
                co = db.query(Company).filter(Company.id == oc.company_id).first()
                company_name = co.name if co else None

    # Last activity: most recent notification or invoice for this supplier
    last_notif = db.query(func.max(SupplierNotification.created_at)).filter(
        SupplierNotification.related_supplier_id == s.id).scalar()
    last_inv = db.query(func.max(SupplierInvoice.created_at)).filter(
        SupplierInvoice.supplier_id == s.id).scalar()
    last_activity = max(filter(None, [last_notif, last_inv, s.created_at]))

    return SupplierResponse(
        id=s.id, name=s.name, email=s.email, nif_cif=s.nif_cif,
        phone=s.phone, address=s.address, iban=_decode_iban(s),
        bank_cert_url=s.bank_cert_url,
        status=s.status.value if s.status else "NEW",
        has_permanent_oc=s.oc_id is not None,
        oc_id=s.oc_id, oc_number=oc_number, company_name=company_name,
        is_active=s.is_active, notes_internal=s.notes_internal,
        gdpr_consent=s.gdpr_consent,
        created_at=s.created_at, updated_at=s.updated_at,
        last_activity=last_activity,
        invoices_count=inv_count, pending_invoices=pending,
    )


@router.get("/{supplier_id}", response_model=SupplierResponse)
async def get_supplier(
    supplier_id: int,
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin_user),
):
    supplier = db.query(Supplier).filter(Supplier.id == supplier_id).first()
    if not supplier:
        raise HTTPException(404, "Supplier not found")

    return _build_supplier_response(supplier, db)


@router.get("/{supplier_id}/bank-cert-url")
async def get_supplier_bank_cert(
    supplier_id: int,
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin_user),
):
    """Generate a signed URL (15min) for the supplier's bank certificate PDF."""
    supplier = db.query(Supplier).filter(Supplier.id == supplier_id).first()
    if not supplier:
        raise HTTPException(404, "Supplier not found")
    if not supplier.bank_cert_url:
        raise HTTPException(404, "No bank certificate uploaded")
    return {"url": get_bank_cert_url(supplier.bank_cert_url)}


@router.get("/{supplier_id}/export-excel")
async def export_supplier_excel(
    supplier_id: int,
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin_user),
):
    """Export supplier invoices as Excel (BytesIO, no filesystem)."""
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from fastapi.responses import StreamingResponse

    supplier = db.query(Supplier).filter(Supplier.id == supplier_id).first()
    if not supplier:
        raise HTTPException(404, "Supplier not found")

    invoices = db.query(SupplierInvoice).filter(
        SupplierInvoice.supplier_id == supplier_id
    ).order_by(desc(SupplierInvoice.created_at)).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Facturas"

    # Header style
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_fill = PatternFill(start_color="27272A", end_color="27272A", fill_type="solid")

    headers = ["Nº Factura", "Fecha", "Proveedor", "OC", "Base", "IVA %", "IVA", "IRPF %", "IRPF", "Total", "Moneda", "Estado"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row, inv in enumerate(invoices, 2):
        ws.cell(row=row, column=1, value=inv.invoice_number)
        ws.cell(row=row, column=2, value=format_date_for_response(inv.date))
        ws.cell(row=row, column=3, value=inv.provider_name)
        ws.cell(row=row, column=4, value=inv.oc_number)
        ws.cell(row=row, column=5, value=inv.base_amount)
        ws.cell(row=row, column=6, value=inv.iva_percentage)
        ws.cell(row=row, column=7, value=inv.iva_amount)
        ws.cell(row=row, column=8, value=inv.irpf_percentage or 0)
        ws.cell(row=row, column=9, value=inv.irpf_amount or 0)
        ws.cell(row=row, column=10, value=inv.final_total)
        ws.cell(row=row, column=11, value=inv.currency or "EUR")
        ws.cell(row=row, column=12, value=inv.status.value if inv.status else "PENDING")

    # Auto-width
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 30)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_name = supplier.name.replace(" ", "_")[:30]
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{safe_name}_facturas.xlsx"'},
    )


@router.put("/{supplier_id}")
async def update_supplier(
    supplier_id: int,
    body: SupplierUpdate,
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin_user),
):
    supplier = db.query(Supplier).filter(Supplier.id == supplier_id).first()
    if not supplier:
        raise HTTPException(404, "Supplier not found")

    ALLOWED_FIELDS = {"name", "nif_cif", "phone", "address", "notes_internal"}
    for field, value in body.model_dump(exclude_unset=True).items():
        if field in ALLOWED_FIELDS:
            setattr(supplier, field, value)

    db.commit()
    return {"message": "Supplier updated"}


@router.put("/{supplier_id}/deactivate")
async def deactivate_supplier(
    supplier_id: int,
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin_user),
):
    """Deactivate supplier and invalidate all their tokens."""
    supplier = db.query(Supplier).filter(Supplier.id == supplier_id).first()
    if not supplier:
        raise HTTPException(404, "Supplier not found")

    supplier.is_active = False
    supplier.status = SupplierStatus.DEACTIVATED
    invalidate_all_supplier_tokens(db, supplier_id)
    db.commit()
    return {"message": f"Supplier {supplier.name} deactivated, all tokens invalidated"}


@router.put("/{supplier_id}/reactivate")
async def reactivate_supplier(
    supplier_id: int,
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin_user),
):
    """Reactivate a previously deactivated supplier."""
    supplier = db.query(Supplier).filter(Supplier.id == supplier_id).first()
    if not supplier:
        raise HTTPException(404, "Supplier not found")

    supplier.is_active = True
    supplier.status = SupplierStatus.ACTIVE
    db.commit()
    return {"message": f"Supplier {supplier.name} reactivated"}


@router.put("/{supplier_id}/assign-oc")
async def assign_oc(
    supplier_id: int,
    body: AssignOCRequest,
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin_user),
):
    """Manually assign an OC to a supplier by OC number."""
    supplier = db.query(Supplier).filter(Supplier.id == supplier_id).first()
    if not supplier:
        raise HTTPException(404, "Supplier not found")

    oc = db.query(SupplierOC).filter(SupplierOC.oc_number == body.oc_number.strip()).first()
    if not oc:
        raise HTTPException(404, f"OC '{body.oc_number}' not found")

    supplier.oc_id = oc.id
    db.commit()
    return {"message": f"OC {oc.oc_number} assigned to {supplier.name}"}


@router.post("/{supplier_id}/notes")
async def add_note(
    supplier_id: int,
    body: NoteRequest,
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin_user),
):
    supplier = db.query(Supplier).filter(Supplier.id == supplier_id).first()
    if not supplier:
        raise HTTPException(404, "Supplier not found")

    timestamp = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M")
    new_note = f"[{timestamp}] {body.note}"
    supplier.notes_internal = (
        f"{supplier.notes_internal}\n{new_note}" if supplier.notes_internal else new_note
    )
    db.commit()
    return {"message": "Note added"}


# ============================================
# INVOICES MANAGEMENT
# ============================================

@router.get("/invoices/all", response_model=List[InvoiceResponse])
async def list_all_invoices(
    status_filter: Optional[str] = Query(None, alias="status"),
    company_id: Optional[int] = None,
    supplier_id: Optional[int] = None,
    oc_number: Optional[str] = None,
    limit: int = Query(50, le=200),
    offset: int = 0,
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin_user),
):
    """List all supplier invoices with filters."""
    query = db.query(SupplierInvoice)

    if supplier_id:
        query = query.filter(SupplierInvoice.supplier_id == supplier_id)
    if status_filter:
        query = query.filter(SupplierInvoice.status == status_filter)
    if company_id:
        query = query.filter(SupplierInvoice.company_id == company_id)
    if oc_number:
        # PERF-M5: LIKE %texto% no usa índice B-tree, pero es aceptable con <1000 facturas.
        # Reconsiderar pg_trgm si el volumen supera 10.000 registros.
        query = query.filter(SupplierInvoice.oc_number.ilike(f"%{oc_number}%"))

    invoices = query.options(
        joinedload(SupplierInvoice.supplier)
    ).order_by(desc(SupplierInvoice.created_at)).offset(offset).limit(limit).all()

    result = []
    for inv in invoices:
        result.append(InvoiceResponse(
            id=inv.id, supplier_id=inv.supplier_id,
            supplier_name=inv.supplier.name if inv.supplier else None,
            invoice_number=inv.invoice_number,
            date=format_date_for_response(inv.date),
            provider_name=inv.provider_name, oc_number=inv.oc_number,
            company_id=inv.company_id,
            base_amount=inv.base_amount, iva_percentage=inv.iva_percentage,
            iva_amount=inv.iva_amount, irpf_percentage=inv.irpf_percentage or 0,
            irpf_amount=inv.irpf_amount or 0, final_total=inv.final_total,
            currency=inv.currency or "EUR", is_foreign=inv.is_foreign or False,
            file_url=inv.file_url if (inv.file_url and inv.file_url.startswith("http")) else get_invoice_pdf_url(inv.file_url) if inv.file_url else "",
            status=inv.status.value if inv.status else "PENDING",
            rejection_reason=inv.rejection_reason, delete_reason=inv.delete_reason,
            created_at=inv.created_at,
        ))

    return result


@router.patch("/invoices/{invoice_id}/assign-oc")
async def assign_invoice_oc(
    invoice_id: int,
    body: AssignInvoiceOCRequest,
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin_user),
):
    """Assign OC to an OC_PENDING invoice, transitioning it to PENDING."""
    invoice = db.query(SupplierInvoice).filter(
        SupplierInvoice.id == invoice_id
    ).first()
    if not invoice:
        raise HTTPException(404, "Invoice not found")

    if invoice.status != InvoiceStatus.OC_PENDING:
        raise HTTPException(400, "Only OC_PENDING invoices can be assigned an OC")

    oc_number = body.oc_number.strip()
    invoice.oc_number = oc_number

    # Try to resolve project_id and company_id from the OC
    project = db.query(Project).filter(
        Project.creative_code.ilike(oc_number)
    ).first()
    if project:
        invoice.project_id = project.id
        invoice.company_id = project.owner_company_id
    else:
        # Try supplier_ocs for company_id
        oc = db.query(SupplierOC).filter(
            SupplierOC.oc_number.ilike(oc_number)
        ).first()
        if oc:
            invoice.company_id = oc.company_id
        else:
            # Free text OC — try to resolve company via prefix map
            from app.services.supplier_ai import resolve_company_from_oc
            invoice.company_id = resolve_company_from_oc(oc_number, db)

    invoice.status = InvoiceStatus.PENDING

    # Notify supplier
    supplier = db.query(Supplier).filter(Supplier.id == invoice.supplier_id).first()
    if supplier:
        _notify(db, NotificationRecipientType.SUPPLIER, supplier.id,
                NotificationEventType.OC_LINKED, "OC Assigned",
                f"Invoice {invoice.invoice_number} linked to OC {oc_number}",
                invoice_id=invoice.id, supplier_id=supplier.id)

    db.commit()
    return {"message": f"OC '{oc_number}' assigned, invoice is now PENDING"}


@router.get("/invoices/{invoice_id}", response_model=InvoiceDetailResponse)
async def get_invoice(
    invoice_id: int,
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin_user),
):
    """Get a single supplier invoice with full details."""
    invoice = db.query(SupplierInvoice).options(
        joinedload(SupplierInvoice.supplier)
    ).filter(SupplierInvoice.id == invoice_id).first()
    if not invoice:
        raise HTTPException(404, "Invoice not found")

    # file_pages not in ORM — fetch via raw SQL
    from sqlalchemy import text as sa_text
    file_pages = None
    try:
        row = db.execute(sa_text("SELECT file_pages FROM supplier_invoices WHERE id = :id"), {"id": invoice_id}).first()
        if row:
            file_pages = row[0]
    except Exception:
        pass

    return InvoiceDetailResponse(
        id=invoice.id, supplier_id=invoice.supplier_id,
        supplier_name=invoice.supplier.name if invoice.supplier else None,
        invoice_number=invoice.invoice_number,
        date=format_date_for_response(invoice.date),
        provider_name=invoice.provider_name, oc_number=invoice.oc_number,
        company_id=invoice.company_id,
        base_amount=invoice.base_amount, iva_percentage=invoice.iva_percentage,
        iva_amount=invoice.iva_amount,
        irpf_percentage=invoice.irpf_percentage or 0,
        irpf_amount=invoice.irpf_amount or 0,
        final_total=invoice.final_total,
        currency=invoice.currency or "EUR", is_foreign=invoice.is_foreign or False,
        file_url=invoice.file_url if (invoice.file_url and invoice.file_url.startswith("http")) else get_invoice_pdf_url(invoice.file_url) if invoice.file_url else "",
        file_pages=file_pages,
        status=invoice.status.value if invoice.status else "PENDING",
        rejection_reason=invoice.rejection_reason, delete_reason=invoice.delete_reason,
        nif_cif=invoice.nif_cif, iban=invoice.iban,
        ia_validation_result=invoice.ia_validation_result,
        created_at=invoice.created_at,
    )


@router.put("/invoices/{invoice_id}/status")
async def update_invoice_status(
    invoice_id: int,
    body: InvoiceStatusUpdate,
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin_user),
):
    """Change invoice status: OC_PENDING → PENDING → APPROVED → PAID."""
    invoice = db.query(SupplierInvoice).filter(SupplierInvoice.id == invoice_id).first()
    if not invoice:
        raise HTTPException(404, "Invoice not found")

    supplier = db.query(Supplier).filter(Supplier.id == invoice.supplier_id).first()
    new_status = body.status.upper()

    # Validate transitions
    valid_transitions = {
        "PENDING": ["APPROVED"],
        "OC_PENDING": ["PENDING"],
        "APPROVED": ["PAID"],
        "DELETE_REQUESTED": [],  # M-17: Must use DELETE endpoint, not status change
    }
    current = invoice.status.value if invoice.status else "PENDING"
    allowed = valid_transitions.get(current, [])

    if new_status not in allowed:
        if current == "DELETE_REQUESTED":
            raise HTTPException(400, "This invoice has a pending deletion request. Use the delete endpoint to confirm it.")
        if current == "PAID":
            raise HTTPException(400, f"Invoice is {current} — no further status changes allowed.")
        raise HTTPException(400, f"Cannot transition from {current} to {new_status}. Allowed: {allowed}")

    invoice.status = InvoiceStatus(new_status)

    # Auto-create ticket in DAZZ Producciones when APPROVED
    if new_status == "APPROVED" and invoice.project_id:
        existing_ticket = db.query(Ticket).filter(
            Ticket.supplier_invoice_id == invoice.id
        ).first()
        if not existing_ticket:
            ticket = Ticket(
                project_id=invoice.project_id,
                date=format_date_for_response(invoice.date),
                provider=invoice.provider_name or "",
                invoice_number=invoice.invoice_number,
                po_notes=invoice.oc_number,
                base_amount=invoice.base_amount,
                iva_percentage=invoice.iva_percentage,
                iva_amount=invoice.iva_amount,
                irpf_percentage=invoice.irpf_percentage or 0,
                irpf_amount=invoice.irpf_amount or 0,
                total_with_iva=invoice.base_amount + invoice.iva_amount,
                final_total=invoice.final_total,
                currency=invoice.currency or "EUR",
                is_foreign=invoice.is_foreign or False,
                invoice_status="Aprobada",
                payment_status="Pendiente",
                type=TicketType.FACTURA,
                file_path=invoice.file_url or "",
                file_name=f"{invoice.invoice_number}.pdf",
                notes="Ticket generated automatically from supplier portal",
                from_supplier_portal=True,
                supplier_id=invoice.supplier_id,
                supplier_invoice_id=invoice.id,
            )
            db.add(ticket)
            logger.info(f"Ticket created in DAZZ for invoice {invoice.invoice_number} -> project {invoice.project_id}")

    # Notifications + emails
    if supplier:
        if new_status == "APPROVED":
            _notify(db, NotificationRecipientType.SUPPLIER, supplier.id,
                    NotificationEventType.APPROVED, "Invoice Approved",
                    f"Invoice {invoice.invoice_number} has been approved",
                    invoice_id=invoice.id, supplier_id=supplier.id)

        elif new_status == "PAID":
            _notify(db, NotificationRecipientType.SUPPLIER, supplier.id,
                    NotificationEventType.PAID, "Invoice Paid",
                    f"Invoice {invoice.invoice_number} — {invoice.final_total:.2f} EUR",
                    invoice_id=invoice.id, supplier_id=supplier.id)
            try:
                send_supplier_invoice_paid(supplier.name, supplier.email,
                                           invoice.invoice_number, invoice.final_total)
            except Exception:
                pass

    # Single commit: status change + ticket + notifications
    db.commit()
    return {"message": f"Invoice status changed to {new_status}"}


@router.delete("/invoices/{invoice_id}")
async def confirm_invoice_deletion(
    invoice_id: int,
    reason: Optional[str] = Body(None, embed=True),
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin_user),
):
    """Delete invoice physically. Admin direct delete or confirm DELETE_REQUESTED."""
    invoice = db.query(SupplierInvoice).filter(SupplierInvoice.id == invoice_id).first()
    if not invoice:
        raise HTTPException(404, "Invoice not found")

    if invoice.status not in (InvoiceStatus.PENDING, InvoiceStatus.OC_PENDING, InvoiceStatus.DELETE_REQUESTED):
        raise HTTPException(400, "Only PENDING, OC_PENDING or DELETE_REQUESTED invoices can be deleted")

    supplier = db.query(Supplier).filter(Supplier.id == invoice.supplier_id).first()

    # Notify supplier only on admin direct delete (not when confirming DELETE_REQUESTED)
    if supplier and invoice.status != InvoiceStatus.DELETE_REQUESTED:
        msg = f"Invoice {invoice.invoice_number} was deleted"
        if reason:
            msg += f": {reason}"
        _notify(db, NotificationRecipientType.SUPPLIER, supplier.id,
                NotificationEventType.DELETED, "Invoice Deleted",
                msg, invoice_id=invoice.id, supplier_id=supplier.id)

    # Cascade: delete or void the linked ticket in DAZZ Producciones
    linked_ticket = db.query(Ticket).filter(
        Ticket.supplier_invoice_id == invoice.id
    ).first()
    if linked_ticket:
        # If ticket has been reviewed (is_reviewed=True), mark as voided instead of deleting
        if linked_ticket.is_reviewed:
            linked_ticket.payment_status = "Anulado"
            linked_ticket.notes = (linked_ticket.notes or "") + "\n[AUTO] Voided: supplier invoice deleted from portal"
            logger.info(f"Ticket {linked_ticket.id} marked as Anulado (had activity)")
        else:
            db.delete(linked_ticket)
            logger.info(f"Ticket {linked_ticket.id} deleted (cascade from supplier invoice)")

    # LOGIC-M5: Delete Cloudinary files with error logging (was silent pass)
    if invoice.file_url:
        try:
            delete_invoice_pdf(invoice.file_url)
        except Exception as e:
            logger.error(f"Failed to delete Cloudinary PDF for invoice {invoice.id}: {e}")

    # Delete page images if they exist (file_pages not in ORM)
    from sqlalchemy import text as sa_text
    try:
        row = db.execute(sa_text("SELECT file_pages FROM supplier_invoices WHERE id = :id"), {"id": invoice.id}).first()
        if row and row[0]:
            import json
            for page_url in json.loads(row[0]):
                try:
                    pid = extract_public_id_from_url(page_url)
                    if pid:
                        import cloudinary.uploader
                        cloudinary.uploader.destroy(pid, resource_type="image")
                except Exception as e:
                    logger.error(f"Failed to delete Cloudinary page image for invoice {invoice.id}, url={page_url}: {e}")
    except Exception as e:
        logger.error(f"Failed to process file_pages for invoice {invoice.id}: {e}")

    db.delete(invoice)
    db.commit()

    return {"message": "Invoice deleted"}


# ============================================
# NOTIFICATIONS
# ============================================

@router.get("/notifications/all", response_model=List[NotificationResponse])
async def list_notifications(
    unread_only: bool = False,
    limit: int = Query(50, le=200),
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin_user),
):
    """Get admin notifications (polling endpoint)."""
    query = db.query(SupplierNotification).filter(
        SupplierNotification.recipient_type == NotificationRecipientType.ADMIN
    )
    if unread_only:
        query = query.filter(SupplierNotification.is_read == False)

    notifications = query.order_by(desc(SupplierNotification.created_at)).limit(limit).all()

    return [NotificationResponse(
        id=n.id, event_type=n.event_type.value if n.event_type else "",
        title=n.title, message=n.message,
        related_invoice_id=n.related_invoice_id,
        related_supplier_id=n.related_supplier_id,
        is_read=n.is_read, created_at=n.created_at,
    ) for n in notifications]


@router.put("/notifications/{notification_id}/read")
async def mark_notification_read(
    notification_id: int,
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin_user),
):
    notif = db.query(SupplierNotification).filter(SupplierNotification.id == notification_id).first()
    if not notif:
        raise HTTPException(404, "Notification not found")

    notif.is_read = True
    db.commit()
    return {"message": "Marked as read"}


@router.put("/notifications/read-all")
async def mark_all_notifications_read(
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin_user),
):
    """Mark all admin notifications as read."""
    db.query(SupplierNotification).filter(
        SupplierNotification.recipient_type == NotificationRecipientType.ADMIN,
        SupplierNotification.is_read == False,
    ).update({"is_read": True})
    db.commit()
    return {"message": "All notifications marked as read"}


# ============================================
# ADMIN UTILITIES
# ============================================

@router.post("/admin/backfill-date-parsed")
async def backfill_date_parsed(
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin_user),
):
    """One-time backfill: parse date string → date_parsed for old invoices with NULL date_parsed."""
    invoices = db.query(SupplierInvoice).filter(
        SupplierInvoice.date_parsed == None,
        SupplierInvoice.date != None,
        SupplierInvoice.date != "",
    ).all()

    updated = 0
    for inv in invoices:
        parsed = parse_invoice_date(inv.date)
        if parsed:
            inv.date_parsed = parsed
            updated += 1

    db.commit()
    return {"message": f"Backfilled {updated}/{len(invoices)} invoices"}


# ============================================
# DASHBOARD
# ============================================

@router.get("/dashboard/stats", response_model=DashboardResponse)
async def dashboard_stats(
    db: Session = Depends(get_db),
    admin: User = Depends(get_current_admin_user),
):
    """KPIs for the supplier dashboard."""
    now = datetime.now(timezone.utc)
    month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    pending = db.query(func.count(SupplierInvoice.id)).filter(
        SupplierInvoice.status == InvoiceStatus.PENDING).scalar() or 0

    approved_month = db.query(func.count(SupplierInvoice.id)).filter(
        SupplierInvoice.status == InvoiceStatus.APPROVED,
        SupplierInvoice.updated_at >= month_start).scalar() or 0

    active = db.query(func.count(Supplier.id)).filter(
        Supplier.is_active == True).scalar() or 0

    paid_month = db.query(func.coalesce(func.sum(SupplierInvoice.final_total), 0.0)).filter(
        SupplierInvoice.status == InvoiceStatus.PAID,
        SupplierInvoice.updated_at >= month_start).scalar() or 0.0

    unread = db.query(func.count(SupplierNotification.id)).filter(
        SupplierNotification.recipient_type == NotificationRecipientType.ADMIN,
        SupplierNotification.is_read == False).scalar() or 0

    return DashboardResponse(
        pending_invoices=pending,
        approved_this_month=approved_month,
        active_suppliers=active,
        total_paid_this_month=float(paid_month),
        unread_notifications=unread,
    )
